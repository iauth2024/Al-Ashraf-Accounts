from itertools import chain, count
import json
from pyexpat.errors import messages
from django import forms
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth import views as auth_views
from django.urls import reverse_lazy
from django.views.generic import TemplateView
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q, Sum
from django.core.exceptions import ValidationError
from django.db import IntegrityError
from datetime import datetime, timedelta, date, timezone
import logging

from .models import HeadOfAccount, Receipt, Voucher
from .forms import ReceiptForm, UploadExcelForm, VoucherFilterForm, VoucherForm, ApprovalForm, RejectionForm, DateRangeForm
from .utils import daterange, generate_auto_receipt_number

###################################################################################################################


from datetime import date, timedelta
from django.shortcuts import render
from django.db.models import Sum
from .models import Receipt, Voucher, UserActivityLog
from .forms import DateRangeForm

def dashboard(request):
    today = date.today()  # Use 'date' from 'datetime'

    form = DateRangeForm(request.GET or None)
    start_date = end_date = None

    if form.is_valid():
        date_range = form.cleaned_data['date_range']

        if date_range == 'today':
            start_date = end_date = today
        elif date_range == 'yesterday':
            start_date = end_date = today - timedelta(days=1)
        # Add more conditions for other date ranges as needed

    # Ensure default date range covers at least one day
    if not start_date or not end_date or start_date > end_date:
        start_date = end_date = today

    # Retrieve user's activities
    user_logs = UserActivityLog.objects.filter(timestamp__range=[start_date, end_date])
    
    # Calculate screen time, work time, logins, receipts created, and vouchers created
    screen_time = user_logs.aggregate(total=Sum('screen_time'))['total'] or 0
    work_time = user_logs.aggregate(total=Sum('work_time'))['total'] or 0
    logins = user_logs.count()
    total_receipts = Receipt.objects.filter(receipt_date__range=[start_date, end_date]).count()
    total_vouchers = Voucher.objects.filter(voucher_date__range=[start_date, end_date]).count()

    context = {
        'form': form,
        'screen_time': screen_time,
        'work_time': work_time,
        'logins': logins,
        'total_receipts': total_receipts,
        'total_vouchers': total_vouchers,
    }

    return render(request, 'dashboard.html', context)


###################################################################################################################
from datetime import datetime, timedelta
from django.contrib.auth.decorators import login_required

@login_required
def homepage(request):
    # Get today's date
    today = datetime.today().date()
    
    # Initialize form and default start/end dates
    form = DateRangeForm(request.GET or None)
    start_date = end_date = None
    
    if form.is_valid():
        date_range = form.cleaned_data['date_range']
        
        # Handle different date ranges
        if date_range == 'today':
            start_date = end_date = today
        elif date_range == 'yesterday':
            start_date = end_date = today - timedelta(days=1)
        elif date_range == 'this_week':
            start_date = today - timedelta(days=today.weekday())
            end_date = today
        elif date_range == 'last_week':
            start_date = today - timedelta(days=today.weekday() + 7)
            end_date = start_date + timedelta(days=6)
        elif date_range == 'this_month':
            start_date = today.replace(day=1)
            end_date = today
        elif date_range == 'last_month':
            first_day_of_this_month = today.replace(day=1)
            last_day_of_last_month = first_day_of_this_month - timedelta(days=1)
            start_date = last_day_of_last_month.replace(day=1)
            end_date = last_day_of_last_month
        elif date_range == 'custom':
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
    
    # Default date range if none provided
    if not start_date or not end_date:
        start_date = today.replace(day=1)
        end_date = today

    # Rest of your logic here


    # Datewise summaries for receipts and vouchers
    receipts_by_date = (
        Receipt.objects.filter(receipt_date__range=[start_date, end_date])
        .values('receipt_date')
        .annotate(total=Count('id'))
        .order_by('-receipt_date')
    )

    vouchers_by_date = (
        Voucher.objects.filter(voucher_date__range=[start_date, end_date])
        .values('voucher_date')
        .annotate(total=Count('id'))
        .order_by('-voucher_date')
    )

    voucher_status_by_date = (
        Voucher.objects.filter(voucher_date__range=[start_date, end_date])
        .values('voucher_date', 'status')
        .annotate(total=Count('id'))
        .order_by('-voucher_date', 'status')
    )

    # Initialize combined data dictionary
    combined_data = {}
    for item in receipts_by_date:
        date = item['receipt_date']
        combined_data[date] = {'receipts': item['total'], 'vouchers': 0, 'statuses': {'approved': 0, 'waiting': 0, 'rejected': 0}}

    for item in vouchers_by_date:
        date = item['voucher_date']
        if date not in combined_data:
            combined_data[date] = {'receipts': 0, 'vouchers': item['total'], 'statuses': {'approved': 0, 'waiting': 0, 'rejected': 0}}
        else:
            combined_data[date]['vouchers'] = item['total']

    for item in voucher_status_by_date:
        date = item['voucher_date']
        status = item['status']
        total = item['total']
        if date not in combined_data:
            combined_data[date] = {'receipts': 0, 'vouchers': 0, 'statuses': {'approved': 0, 'waiting': 0, 'rejected': 0}}
        combined_data[date]['statuses'][status] = total

    context = {
        'total_receipts': Receipt.objects.count(),
        'total_vouchers': Voucher.objects.count(),
        'combined_data': combined_data,
        'form': form,
    }

    return render(request, 'homepage.html', context)

from datetime import datetime


###################################################################################################################
from django.shortcuts import render

def custom_404_view(request, exception):
    return render(request, '404.html', status=404)
def placeholder_view(request):
    return HttpResponse("This is a placeholder view")
from django.http import HttpResponse

def test_view(request):
    return HttpResponse("Test view works!")

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('homepage')  # Redirect to homepage after login
        else:
            # Handle invalid login
            return render(request, 'login.html', {'error_message': 'Invalid username or password'})
    else:
        return render(request, 'login.html')

def logout_view(request):
    if request.method == 'POST':
        logout(request)
        return redirect('login')  # Redirect to login page after logout
    else:
        # Handle GET requests by redirecting to the login page
        return redirect('login')

###################################################################################################################





def create_receipt(request):
    if request.method == 'POST':
        form = ReceiptForm(request.POST)
        if form.is_valid():
            receipt = form.save(commit=False)
            # Manual entry for all payments, no auto-generation
            receipt.manual_receipt_no = form.cleaned_data.get('manual_receipt_no')
            receipt.save()
            messages.success(request, 'Receipt created successfully!')
            return redirect('receipt_success', receipt_id=receipt.id)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = ReceiptForm(initial={'mode_of_payment': 'Cash'})
    
    return render(request, 'create_receipt.html', {'form': form})












def receipt_success(request, receipt_id):
    receipt = get_object_or_404(Receipt, id=receipt_id)
    return render(request, 'receipt_success.html', {'receipt': receipt})
###################################################################################################################
import datetime
from django.shortcuts import render
from django.http import HttpResponse
from .models import Receipt
import openpyxl
from io import BytesIO
from reportlab.lib.pagesizes import A4 # type: ignore
from reportlab.lib import colors # type: ignore
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph # type: ignore
from reportlab.lib.units import inch # type: ignore
from django.contrib.auth.decorators import login_required

@login_required
def list_receipts(request):
    receipts = Receipt.objects.all()

    # Apply search filters
    manual_book_no = request.GET.get('manual_book_no')
    if manual_book_no:
        receipts = receipts.filter(manual_book_no__icontains=manual_book_no)

    manual_receipt_no = request.GET.get('manual_receipt_no')
    if manual_receipt_no:
        receipts = receipts.filter(manual_receipt_no__icontains=manual_receipt_no)

    name = request.GET.get('name')
    if name:
        receipts = receipts.filter(name__icontains=name)

    phone = request.GET.get('phone')
    if phone:
        receipts = receipts.filter(phone__icontains=phone)

    type_of_receipt = request.GET.get('type_of_receipt')
    if type_of_receipt:
        receipts = receipts.filter(type_of_receipt__icontains=type_of_receipt)

    mode_of_payment = request.GET.get('mode_of_payment')
    if mode_of_payment:
        receipts = receipts.filter(mode_of_payment__icontains=mode_of_payment)

    transaction_id = request.GET.get('transaction_id')
    if transaction_id:
        receipts = receipts.filter(transaction_id__icontains=transaction_id)

    cheque_number = request.GET.get('cheque_number')
    if cheque_number:
        receipts = receipts.filter(cheque_number__icontains=cheque_number)

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        receipts = receipts.filter(receipt_date__range=[start_date, end_date])

    # Pagination
    paginator = Paginator(receipts, 10)  # Show 10 receipts per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'manual_book_no': manual_book_no,
        'manual_receipt_no': manual_receipt_no,
        'name': name,
        'phone': phone,
        'type_of_receipt': type_of_receipt,
        'mode_of_payment': mode_of_payment,
        'transaction_id': transaction_id,
        'cheque_number': cheque_number,
        'start_date': start_date,
        'end_date': end_date,
    }

    return render(request, 'list_receipts.html', context)

def download_receipts_excel(receipts):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=receipts.xlsx'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Receipts'

    headers = ['Manual Book No', 'Manual Receipt No', 'Name', 'Phone', 'Type of Receipt', 'Mode of Payment', 'Transaction ID', 'Cheque Number', 'Amount', 'Receipt Date']
    worksheet.append(headers)

    for receipt in receipts:
        row = [
            receipt.manual_book_no,
            receipt.manual_receipt_no,
            receipt.name,
            receipt.phone,
            receipt.type_of_receipt,
            receipt.mode_of_payment,
            receipt.transaction_id,
            receipt.cheque_number,
            receipt.amount,
            receipt.receipt_date.strftime('%Y-%m-%d')
        ]
        worksheet.append(row)

    workbook.save(response)
    return response

def download_receipts_pdf(receipts):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=receipts.pdf'

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=10, leftMargin=10, topMargin=20, bottomMargin=20)

    elements = []
    data = [
        ['Manual B.No', 'Manual R.No', 'Name', 'Phone', 'Receipt type', 'Payment type', 'Txn ID', 'Cheque No', 'Amount', 'Receipt Date']
    ]

    # Set up a style for paragraphs
    styles = getSampleStyleSheet() # type: ignore
    styleN = styles['BodyText']
    styleN.wordWrap = 'CJK'  # Enable wrapping

    for receipt in receipts:
        row = [
            Paragraph(str(receipt.manual_book_no) if receipt.manual_book_no else '', styleN),
            Paragraph(str(receipt.manual_receipt_no) if receipt.manual_receipt_no else '', styleN),
            Paragraph(str(receipt.name) if receipt.name else '', styleN),
            Paragraph(str(receipt.phone) if receipt.phone else '', styleN),
            Paragraph(str(receipt.type_of_receipt) if receipt.type_of_receipt else '', styleN),
            Paragraph(str(receipt.mode_of_payment) if receipt.mode_of_payment else '', styleN),
            Paragraph(str(receipt.transaction_id) if receipt.transaction_id else '', styleN),
            Paragraph(str(receipt.cheque_number) if receipt.cheque_number else '', styleN),
            Paragraph(str(receipt.amount) if receipt.amount else '', styleN),
            Paragraph(receipt.receipt_date.strftime('%Y-%m-%d') if receipt.receipt_date else '', styleN)
        ]
        data.append(row)

    table = Table(data, colWidths=[0.75 * inch] * len(data[0]))
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table.setStyle(style)
    elements.append(table)
    doc.build(elements)

    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response


###################################################################################################################
@login_required
def convert_amount_to_words(request):
    amount = request.GET.get('amount')
    if amount:
        try:
            amount = float(amount)
            amount_in_words = num2words(amount) # type: ignore
            return JsonResponse({'amount_in_words': amount_in_words})
        except ValueError:
            return JsonResponse({'error': 'Invalid amount'}, status=400)
    return JsonResponse({'error': 'No amount provided'}, status=400)

@login_required
@staff_member_required
def approver_page(request):
    # Get search and filter parameters from GET request
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    entries_per_page = int(request.GET.get('entries_per_page', 10))

    # Build the query based on search and filter parameters
    vouchers = Voucher.objects.all()
    if search_query:
        vouchers = vouchers.filter(
            Q(voucher_no__icontains=search_query) |
            Q(paid_to__icontains=search_query) |
            Q(payment_purpose__icontains=search_query)
        )
    if status_filter:
        vouchers = vouchers.filter(status=status_filter)

    # Sort vouchers based on status
    sorted_vouchers = sorted(vouchers, key=lambda v: (
        v.status == 'approved',
        v.status == 'rejected',
        v.status not in ['waiting', 'waiting for approval']
    ))

    # Paginate sorted vouchers
    paginator = Paginator(sorted_vouchers, entries_per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'approver_page.html', {
        'page_obj': page_obj,
        'entries_per_page': entries_per_page,
        'search_query': search_query,
        'status_filter': status_filter
    })

from django.db import transaction
from django.urls import reverse
from django.http import HttpResponseRedirect

@login_required
def create_voucher(request):
    if request.method == 'POST':
        form = VoucherForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():  # Ensure atomicity
                    voucher = form.save(commit=False)
                    voucher.created_by = request.user
                    voucher.save()
                messages.success(request, 'Voucher created successfully!')
                return HttpResponseRedirect(reverse('voucher_success', args=[voucher.id]))
            except Exception as e:
                messages.error(request, f"An error occurred: {str(e)}")
        else:
            # Collect form errors for messaging framework (optional)
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = VoucherForm()

    return render(request, 'create_voucher.html', {'form': form})

@login_required
def voucher_success(request, voucher_id):
    voucher = get_object_or_404(Voucher, id=voucher_id)
    return render(request, 'voucher_success.html', {'voucher': voucher})



@login_required
def edit_voucher(request, voucher_id):
    voucher = get_object_or_404(Voucher, id=voucher_id)
    
    if request.method == 'POST':
        form = VoucherForm(request.POST, instance=voucher)
        if form.is_valid():
            updated_voucher = form.save(commit=False)
            if any(field in form.changed_data for field in form.fields if field != 'status'):
                updated_voucher.status = 'waiting for approval'
            updated_voucher.save()
            return redirect('view_voucher', voucher_id=voucher.id)
    else:
        form = VoucherForm(instance=voucher)
    
    return render(request, 'edit_voucher.html', {'form': form, 'voucher': voucher})



from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponseForbidden, JsonResponse

@login_required
@user_passes_test(lambda u: u.is_staff)
def view_voucher(request, voucher_id):
    voucher = get_object_or_404(Voucher, id=voucher_id)

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'approve':
            voucher.status = 'approved'
            voucher.approved_by = request.user
            voucher.save()
        elif action == 'reject':
            rejection_reason = request.POST.get('rejection_reason', '')
            if rejection_reason.strip() == '':
                return HttpResponseForbidden("Rejection reason is required.")
            voucher.status = 'rejected'
            voucher.rejection_reason = rejection_reason
            voucher.save()
        return redirect('view_voucher', voucher_id=voucher.id)

    return render(request, 'view_voucher.html', {'voucher': voucher})

@login_required
def edit_vouchers_list(request):
    vouchers = Voucher.objects.all()
    return render(request, 'edit_vouchers_list.html', {'vouchers': vouchers})



@login_required
@staff_member_required
def approve_voucher(request, voucher_id):
    voucher = get_object_or_404(Voucher, id=voucher_id)

    if request.method == 'POST':
        form = ApprovalForm(request.POST, instance=voucher, user=request.user)
        if form.is_valid():
            voucher = form.save(commit=False)
            if 'reject' in request.POST:
                voucher.status = 'rejected'
            else:
                voucher.status = 'approved'
                voucher.approved_by = request.user
            voucher.save()
            return redirect('view_voucher', voucher_id=voucher.id)
    else:
        form = ApprovalForm(instance=voucher, user=request.user)

    return render(request, 'approve_voucher.html', {'form': form, 'voucher': voucher})



@login_required
@staff_member_required
def reject_voucher(request, voucher_id):
    voucher = get_object_or_404(Voucher, id=voucher_id)
    
    if request.method == 'POST':
        form = RejectionForm(request.POST, instance=voucher)
        if form.is_valid():
            voucher = form.save(commit=False)
            voucher.status = 'rejected'
            voucher.save()
            return redirect('view_voucher', voucher_id=voucher_id)
    else:
        form = RejectionForm(instance=voucher)
    
    return render(request, 'reject_voucher.html', {'form': form, 'voucher': voucher})
from django.core.paginator import Paginator
###############################################################################################################################
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.db.models import Q
from django.core.paginator import Paginator
from .models import Voucher
import csv
import xlwt
from io import BytesIO
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter # type: ignore
from reportlab.pdfgen import canvas # type: ignore

@login_required
def list_vouchers(request):
    form = VoucherFilterForm(request.GET or None)
    vouchers = Voucher.objects.all()

    # Apply filters only if the form is valid
    if form.is_valid():
        # Filtering by voucher_no
        if form.cleaned_data.get('voucher_no'):
            vouchers = vouchers.filter(voucher_no__icontains=form.cleaned_data['voucher_no'])
        
        # Filtering by paid_to
        if form.cleaned_data.get('paid_to'):
            vouchers = vouchers.filter(paid_to__icontains=form.cleaned_data['paid_to'])

        # Filtering by on_account_of
        if form.cleaned_data.get('on_account_of'):
            vouchers = vouchers.filter(on_account_of__icontains=form.cleaned_data['on_account_of'])

        # Filtering by head_of_account
        if form.cleaned_data.get('head_of_account'):
            vouchers = vouchers.filter(head_of_account__name__icontains=form.cleaned_data['head_of_account'])

        # Filtering by mode_of_payment
        if form.cleaned_data.get('mode_of_payment'):
            vouchers = vouchers.filter(mode_of_payment=form.cleaned_data['mode_of_payment'])

        # Filtering by transaction_id
        if form.cleaned_data.get('transaction_id'):
            vouchers = vouchers.filter(transaction_id__icontains=form.cleaned_data['transaction_id'])

        # Filtering by amount (min and max)
        if form.cleaned_data.get('amount_min') is not None:
            vouchers = vouchers.filter(amount__gte=form.cleaned_data['amount_min'])
        if form.cleaned_data.get('amount_max') is not None:
            vouchers = vouchers.filter(amount__lte=form.cleaned_data['amount_max'])

        # Filtering by voucher_date (start_date and end_date)
        if form.cleaned_data.get('start_date'):
            vouchers = vouchers.filter(voucher_date__gte=form.cleaned_data['start_date'])
        if form.cleaned_data.get('end_date'):
            vouchers = vouchers.filter(voucher_date__lte=form.cleaned_data['end_date'])

        # Filtering by received_by
        if form.cleaned_data.get('received_by'):
            vouchers = vouchers.filter(received_by__icontains=form.cleaned_data['received_by'])

        # Filtering by status
        if form.cleaned_data.get('status'):
            vouchers = vouchers.filter(status=form.cleaned_data['status'])

    # Apply pagination AFTER filtering
    paginator = Paginator(vouchers, 10)  # Show 10 vouchers per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Build the query parameters to maintain filters during pagination
    query_params = request.GET.copy()
    if 'page' in query_params:
        query_params.pop('page')  # Remove 'page' parameter to avoid duplication in URLs

    return render(request, 'list_vouchers.html', {
        'form': form,
        'page_obj': page_obj,
        'query_params': query_params,  # Pass query params for pagination links
    })

def download_vouchers(vouchers, download_format):
    if download_format == 'excel':
        return download_excel(vouchers)
    elif download_format == 'pdf':
        return download_pdf(vouchers)
    else:
        return HttpResponse(status=400)
from django.http import HttpResponse
from openpyxl import Workbook
from io import BytesIO
from .models import Voucher




def download_excel(request):
    # Queryset for all vouchers
    vouchers = Voucher.objects.all()

    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = 'Vouchers'

    # Define the columns, including 'Approved By'
    columns = [
        'Voucher No', 'Paid To', 'On Account Of', 'Head Of Account',
        'Mode Of Payment', 'Transaction ID', 'Amount', 'Date', 'Received By', 'Status', 'Approved By'
    ]

    # Write the column headers to the worksheet
    for col_num, column_title in enumerate(columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)

    # Write the rows to the worksheet
    rows = vouchers.values_list(
        'voucher_no', 'paid_to', 'on_account_of', 'head_of_account__name',
        'mode_of_payment', 'transaction_id', 'amount', 'voucher_date',
        'received_by', 'status', 'approved_by__username'  # Change to username or use other attribute
    )
    for row_num, row in enumerate(rows, 2):
        for col_num, cell_value in enumerate(row, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    # Save the workbook to a BytesIO buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Set the HTTP response content type to Excel format
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="vouchers.xlsx"'
    return response


from reportlab.lib.pagesizes import A4, landscape # type: ignore



def download_pdf(request):
    # Queryset for all vouchers
    vouchers = Voucher.objects.all()

    # Create a PDF buffer
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))  # Set pagesize to landscape A4
    elements = []

    # Define the columns, including 'Approved By'
    columns = [
        'Voucher No', 'Paid To', 'On Account Of', 'Head Of Account',
        'Mode Of Payment', 'Transaction ID', 'Amount', 'Date', 'Received By', 'Status', 'Approved By'
    ]

    # Fetch the rows from the queryset
    rows = vouchers.values_list(
        'voucher_no', 'paid_to', 'on_account_of', 'head_of_account__name',
        'mode_of_payment', 'transaction_id', 'amount', 'voucher_date',
        'received_by', 'status', 'approved_by__username'  # Change to username or use other attribute
    )

    # Create a table data structure for PDF
    data = [columns] + [list(row) for row in rows]

    # Create the table and style it
    table = Table(data)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 0), (-1, -1), 8),  # Adjust font size to fit more content
        ('ALIGN', (0, 1), (-1, -1), 'LEFT')  # Align text to the left for readability
    ])
    table.setStyle(style)
    elements.append(table)

    # Build the PDF document
    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="vouchers.pdf"'
    return response


#####################################################################################################################
def delete_voucher(request, voucher_id):
    voucher = get_object_or_404(Voucher, id=voucher_id)
    voucher.delete()
    # Redirect to a success page or another appropriate URL
    return redirect('list_vouchers')  # Assuming 'list_vouchers' is the name of your list view

###################################################################################################################

class CustomPasswordResetView(auth_views.PasswordResetView):
    template_name = 'custom_password_reset.html'
    email_template_name = 'custom_password_reset_email.html'
    success_url = reverse_lazy('password_reset_done')

class CustomPasswordResetDoneView(auth_views.PasswordResetDoneView):
    template_name = 'custom_password_reset_done.html'

class CustomPasswordResetConfirmView(auth_views.PasswordResetConfirmView):
    template_name = 'custom_password_reset_confirm.html'
    success_url = reverse_lazy('password_reset_complete')

class CustomPasswordResetCompleteView(auth_views.PasswordResetCompleteView):
    template_name = 'custom_password_reset_complete.html'



####################################################################################################################



################################################################################



class DateRangeForm(forms.Form):
    DATE_RANGE_CHOICES = [
        ('today', 'Today'),
        ('yesterday', 'Yesterday'),
        ('this_week', 'This Week'),
        ('last_week', 'Last Week'),
        ('this_month', 'This Month'),
        ('last_month', 'Last Month'),
        ('custom', 'Custom'),
    ]
    
    date_range = forms.ChoiceField(choices=DATE_RANGE_CHOICES)
    start_date = forms.DateField(required=False, widget=forms.TextInput(attrs={'type': 'date'}))
    end_date = forms.DateField(required=False, widget=forms.TextInput(attrs={'type': 'date'}))

def get_date_range(date_range):
    today = timezone.now().date()
    if date_range == 'today':
        start_date = end_date = today
    elif date_range == 'yesterday':
        start_date = end_date = today - timedelta(days=1)
    elif date_range == 'this_week':
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    elif date_range == 'last_week':
        start_date = today - timedelta(days=today.weekday() + 7)
        end_date = start_date + timedelta(days=6)
    elif date_range == 'this_month':
        start_date = today.replace(day=1)
        end_date = today
    elif date_range == 'last_month':
        first_day_of_current_month = today.replace(day=1)
        last_day_of_last_month = first_day_of_current_month - timedelta(days=1)
        start_date = last_day_of_last_month.replace(day=1)
        end_date = last_day_of_last_month
    else:
        start_date = end_date = today
    return start_date, end_date

def ledger_page(request):
    form = DateRangeForm(request.GET or None)

    # Get the current date
    today = timezone.now().date()

    # Set default start date as the start of the financial year (assuming April 1)
    financial_year_start = date(today.year if today.month >= 4 else today.year - 1, 4, 1)

    # Get date range from the form or set default
    if form.is_valid():
        date_range = form.cleaned_data['date_range']
        start_date, end_date = get_date_range(date_range)

        if date_range == 'custom':
            start_date = form.cleaned_data['start_date'] or financial_year_start
            end_date = form.cleaned_data['end_date'] or today
    else:
        start_date = financial_year_start
        end_date = today

    # Debugging output
    print(f"Start Date: {start_date}, End Date: {end_date}")

    # Fetch transactions within the date range
    receipts = Receipt.objects.filter(receipt_date__range=[start_date, end_date])
    vouchers = Voucher.objects.filter(voucher_date__range=[start_date, end_date])

    # Debugging output
    print(f"Receipts Count: {receipts.count()}, Vouchers Count: {vouchers.count()}")

    # Combine receipts and vouchers into a single list
    transactions = sorted(
        chain(receipts, vouchers),
        key=lambda x: x.receipt_date if hasattr(x, 'receipt_date') else x.voucher_date
    )

    # Paginate the transactions list
    paginator = Paginator(transactions, 10)  # Show 10 transactions per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'form': form,
        'page_obj': page_obj,
        'start_date': start_date,
        'end_date': end_date,
        'today': today,
    }
    return render(request, 'ledger.html', context)
 ###############################################################################################################


from django.db.models import Count, Sum
from django.shortcuts import render
from .models import Receipt, Voucher
from .forms import DateRangeForm
from django.utils import timezone
from datetime import timedelta
from django.db.models import Sum, Count, F
from django.utils import timezone

def get_date_range(date_range):
    today = timezone.now().date()
    if date_range == 'today':
        start_date = end_date = today
    elif date_range == 'yesterday':
        start_date = end_date = today - timedelta(days=1)
    elif date_range == 'this_week':
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    elif date_range == 'last_week':
        start_date = today - timedelta(days=today.weekday() + 7)
        end_date = start_date + timedelta(days=6)
    elif date_range == 'this_month':
        start_date = today.replace(day=1)
        end_date = today
    elif date_range == 'last_month':
        first_day_of_current_month = today.replace(day=1)
        last_day_of_last_month = first_day_of_current_month - timedelta(days=1)
        start_date = last_day_of_last_month.replace(day=1)
        end_date = last_day_of_last_month
    else:
        start_date = end_date = today
    return start_date, end_date


@login_required
def ledger_page_details(request):
    # Instantiate the date form
    date_form = DateRangeForm(request.GET or None)

    # Initialize date range
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Filter data based on the date range if provided
    receipts_query = Receipt.objects.all()
    vouchers_query = Voucher.objects.all()

    if start_date and end_date:
        receipts_query = receipts_query.filter(receipt_date__range=[start_date, end_date])
        vouchers_query = vouchers_query.filter(voucher_date__range=[start_date, end_date])

    # Calculate total amounts for receipts and vouchers
    total_receipts_amount = receipts_query.aggregate(Sum('amount'))['amount__sum'] or 1  # Avoid division by zero
    total_vouchers_amount = vouchers_query.aggregate(Sum('amount'))['amount__sum'] or 1

    # Summarize receipts
    receipt_summary = receipts_query.values('type_of_receipt').annotate(
        count=Count('id'),
        total_amount=Sum('amount'),
        percentage=Sum('amount') * 100 / total_receipts_amount
    )

    # Summarize vouchers
    voucher_summary = vouchers_query.values('head_of_account__name').annotate(
        count=Count('id'),
        total_amount=Sum('amount'),
        percentage=Sum('amount') * 100 / total_vouchers_amount
    )

    context = {
        'date_form': date_form,
        'receipt_summary': receipt_summary,
        'voucher_summary': voucher_summary,
    }

    return render(request, 'ledger_page_details.html', context)





def receipt_detail(request, type_of_receipt):
    date_form = DateRangeForm(request.GET or None)
    date_range = date_form.cleaned_data if date_form.is_valid() else {}
    
    # Filter Receipts based on type and date range
    receipts = Receipt.objects.filter(type_of_receipt=type_of_receipt)
    if date_range:
        start_date, end_date = get_date_range(date_range.get('date_range', 'today'))
        receipts = receipts.filter(receipt_date__range=[start_date, end_date])
    
    # Debugging output
    print(receipts.values())  # Add this line to inspect the receipt objects
    
    context = {
        'receipts': receipts,
        'type_of_receipt': type_of_receipt,
        'date_form': date_form
    }
    return render(request, 'receipt_detail.html', context)



def voucher_detail(request, head_of_account):
    date_form = DateRangeForm(request.GET or None)
    date_range = date_form.cleaned_data if date_form.is_valid() else {}
    
    # Filter Vouchers based on head_of_account and date range
    vouchers = Voucher.objects.filter(head_of_account__name=head_of_account)
    if date_range:
        start_date, end_date = get_date_range(date_range.get('date_range', 'today'))
        vouchers = vouchers.filter(voucher_date__range=[start_date, end_date])
    
    context = {
        'vouchers': vouchers,
        'head_of_account': head_of_account,
        'date_form': date_form
    }
    return render(request, 'voucher_detail.html', context)






###############################################################################################################


from django.shortcuts import render
from django.db.models import Sum
from datetime import date
from .models import Receipt, Voucher

from datetime import date, datetime

def trail_balance(request):
    # Get the date range from the request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    financial_year = request.GET.get('financial_year')

    # Default to the financial year's start if no dates provided
    if not start_date:
        if financial_year and financial_year.isdigit():
            start_date = f"{financial_year}-04-01"
        else:
            start_date = date.today().replace(day=1, month=4).strftime('%Y-%m-%d')
    if not end_date:
        end_date = date.today().strftime('%Y-%m-%d')

    try:
        start_date = date.fromisoformat(start_date)
        end_date = date.fromisoformat(end_date)
    except ValueError:
        start_date = date.today().replace(day=1, month=4)
        end_date = date.today()

    # Adjust end_date to include the entire day (if using DateTimeField)
    end_date = datetime.combine(end_date, datetime.max.time())
    start_date = datetime.combine(start_date, datetime.min.time())

    # Filter Receipts and Vouchers
    receipts = Receipt.objects.filter(receipt_date__range=(start_date, end_date))
    vouchers = Voucher.objects.filter(voucher_date__range=(start_date, end_date))

    # Calculate totals
    total_receipts = receipts.aggregate(total_amount=Sum('amount'))['total_amount'] or 0
    total_vouchers = vouchers.aggregate(total_amount=Sum('amount'))['total_amount'] or 0

    # Group by type or head of account
    receipt_groups = receipts.values('type_of_receipt').annotate(total_amount=Sum('amount'))
    voucher_groups = vouchers.values('head_of_account__name').annotate(total_amount=Sum('amount'))

    # Prepare context
    context = {
        'start_date': start_date.date(),
        'end_date': end_date.date(),
        'financial_year': financial_year,
        'total_receipts': total_receipts,
        'total_vouchers': total_vouchers,
        'receipt_groups': receipt_groups,
        'voucher_groups': voucher_groups,
    }

    return render(request, 'trail_balance.html', context)


################################################################################################################################
from datetime import datetime, timedelta
from django.db.models import Sum
from django.utils import timezone
from django.shortcuts import render
from .models import Receipt, Voucher, Contra

def day_book(request):
    # Date range filter
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Default to today if no date range is provided
    if not start_date or not end_date:
        today = timezone.now().date()
        start_date = today
        end_date = today
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

    # Calculate Previous Day for Opening Balance
    previous_day = start_date - timedelta(days=1)

    # Calculate opening balance as of the previous day
    total_cash_receipts = Receipt.objects.filter(
        mode_of_payment__iexact='cash',
        receipt_date__lte=previous_day
    ).aggregate(total=Sum('amount'))['total'] or 0

    total_bank_receipts = Receipt.objects.filter(
        mode_of_payment__in=['UPI', 'Bank Transfer', 'Cheque', 'bank_transfer', 'cheque'],
        receipt_date__lte=previous_day
    ).aggregate(total=Sum('amount'))['total'] or 0

    total_cash_payments = Voucher.objects.filter(
        mode_of_payment__iexact='cash',
        voucher_date__lte=previous_day
    ).aggregate(total=Sum('amount'))['total'] or 0

    total_bank_payments = Voucher.objects.filter(
        mode_of_payment__in=['UPI', 'Bank Transfer', 'Cheque', 'bank_transfer', 'cheque'],
        voucher_date__lte=previous_day
    ).aggregate(total=Sum('amount'))['total'] or 0

    # Adjusted contra transactions calculation (previous day)
    contra_deposit = Contra.objects.filter(
        contra_type='deposit',
        date__lte=previous_day
    ).aggregate(total=Sum('amount'))['total'] or 0

    contra_withdraw = Contra.objects.filter(
        contra_type='withdraw',
        date__lte=previous_day
    ).aggregate(total=Sum('amount'))['total'] or 0

    # Calculate Opening Balances
    opening_cash_balance = total_cash_receipts - total_cash_payments + contra_withdraw - contra_deposit
    opening_bank_balance = total_bank_receipts - total_bank_payments + contra_deposit - contra_withdraw
    opening_total_balance = opening_cash_balance + opening_bank_balance

    # Calculate Receipts and Payments within the date range
    receipts = Receipt.objects.filter(receipt_date__range=(start_date, end_date))
    vouchers = Voucher.objects.filter(voucher_date__range=(start_date, end_date))

    # Calculate totals for receipts and vouchers
    total_cash_receipts = receipts.filter(mode_of_payment__iexact='cash').aggregate(total=Sum('amount'))['total'] or 0
    total_bank_receipts = receipts.filter(mode_of_payment__in=['UPI', 'Bank Transfer', 'Cheque', 'bank_transfer', 'cheque']).aggregate(total=Sum('amount'))['total'] or 0
    total_cash_payments = vouchers.filter(mode_of_payment__iexact='cash').aggregate(total=Sum('amount'))['total'] or 0
    total_bank_payments = vouchers.filter(mode_of_payment__in=['UPI', 'Bank Transfer', 'Cheque', 'bank_transfer', 'cheque']).aggregate(total=Sum('amount'))['total'] or 0

    # Calculate Contras within the date range
    contra_deposit = Contra.objects.filter(
        contra_type='deposit',
        date__range=(start_date, end_date)
    ).aggregate(total=Sum('amount'))['total'] or 0

    contra_withdraw = Contra.objects.filter(
        contra_type='withdraw',
        date__range=(start_date, end_date)
    ).aggregate(total=Sum('amount'))['total'] or 0

    # Calculate Closing Balances
    closing_cash_balance = opening_cash_balance + total_cash_receipts - total_cash_payments - contra_deposit + contra_withdraw
    closing_bank_balance = opening_bank_balance + total_bank_receipts - total_bank_payments + contra_deposit - contra_withdraw
    closing_total_balance = closing_cash_balance + closing_bank_balance

    # Pass all values to the template
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'opening_cash_balance': opening_cash_balance,
        'opening_bank_balance': opening_bank_balance,
        'opening_total_balance': opening_total_balance,
        'cash_receipts': total_cash_receipts,
        'bank_receipts': total_bank_receipts,
        'cash_payments': total_cash_payments,
        'bank_payments': total_bank_payments,
        'contra_deposit': contra_deposit,
        'contra_withdraw': contra_withdraw,
        'closing_cash_balance': closing_cash_balance,
        'closing_bank_balance': closing_bank_balance,
        'closing_total_balance': closing_total_balance,
    }

    return render(request, 'day_book.html', context)





###############################################################################################################



from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .forms import ContraForm
from .models import Contra, Balance
from django.utils import timezone


from django.shortcuts import render, redirect
from .forms import ContraForm
from .models import Balance

from django.shortcuts import render, redirect
from django.utils import timezone
from .forms import ContraForm
from .models import Balance

def create_contra(request):
    if request.method == 'POST':
        form = ContraForm(request.POST)
        if form.is_valid():
            contra = form.save(commit=False)
            contra.performed_by = request.user
            contra.date = form.cleaned_data['date']
            
            # Fetch current balances for the contra transaction date
            balances = fetch_balances(contra.date, contra.date)

            if contra.contra_type == 'withdraw':
                if contra.amount <= balances['closing_balance_bank']:
                    # Update balances for bank withdrawal and cash receipt
                    balances['closing_balance_bank'] -= contra.amount  # Deduct from bank balance
                    balances['total_receipts_cash'] += contra.amount  # Add to cash receipts
                else:
                    return render(request, 'create_contra.html', {'form': form, 'error': 'Insufficient bank balance.'})

            elif contra.contra_type == 'deposit':
                if contra.amount <= balances['closing_balance_cash']:
                    # Update balances for cash deposit and bank receipt
                    balances['closing_balance_cash'] -= contra.amount  # Deduct from cash balance
                    balances['total_receipts_bank'] += contra.amount  # Add to bank receipts
                else:
                    return render(request, 'create_contra.html', {'form': form, 'error': 'Insufficient cash balance.'})

            # Update the balance records
            balance, created = Balance.objects.get_or_create()
            balance.cash_balance = balances['closing_balance_cash']  # Update closing cash balance
            balance.non_cash_balance = balances['closing_balance_bank']  # Update closing bank balance
            balance.total_receipts_cash = balances['total_receipts_cash']  # Update total cash receipts
            balance.total_receipts_bank = balances['total_receipts_bank']  # Update total bank receipts
            balance.save()

            contra.save()
            return redirect('list_contra')
    else:
        form = ContraForm()

    context = {
        'form': form,
        'today': timezone.now().date()
    }
    return render(request, 'create_contra.html', context)


def delete_contra(request, contra_no):
    contra = get_object_or_404(Contra, pk=contra_no)
    if request.method == 'POST':
        # Fetch balances before reversing the contra transaction
        balances = fetch_balances(contra.date, contra.date)
        
        if contra.contra_type == 'withdraw':
            # Revert the changes made by the withdrawal contra
            balances['closing_balance_bank'] += contra.amount  # Restore bank balance
            balances['total_receipts_cash'] -= contra.amount  # Revert cash receipts
        elif contra.contra_type == 'deposit':
            # Revert the changes made by the deposit contra
            balances['closing_balance_cash'] += contra.amount  # Restore cash balance
            balances['total_receipts_bank'] -= contra.amount  # Revert bank receipts

        # Update the balance records
        balance, created = Balance.objects.get_or_create()
        balance.cash_balance = balances['closing_balance_cash']  # Update closing cash balance
        balance.non_cash_balance = balances['closing_balance_bank']  # Update closing bank balance
        balance.save()

        # Delete the contra entry
        contra.delete()
        return redirect('list_contra')

    return render(request, 'delete_contra.html', {'contra': contra})


def fetch_balances(start_date, end_date):
    """
    Fetches the balances based on the most recent Balance record.
    Returns a dictionary with cash and bank balances, along with total receipts.
    """
    balance = Balance.objects.first()  # Get the first balance record (most recent)
    
    if balance:
        return {
            'closing_balance_cash': balance.cash_balance,
            'closing_balance_bank': balance.non_cash_balance,
            'total_receipts_cash': balance.total_receipts_cash,
            'total_receipts_bank': balance.total_receipts_bank,
        }
    else:
        # Default balances when no records exist
        return {
            'closing_balance_cash': 0,
            'closing_balance_bank': 0,
            'total_receipts_cash': 0,
            'total_receipts_bank': 0,
        }




from .forms import ContraFilterForm

@login_required
def list_contra(request):
    # Initialize filter form
    form = ContraFilterForm(request.GET or None)
    contras = Contra.objects.all()

    if form.is_valid():
        contra_no = form.cleaned_data.get('contra_no')
        date = form.cleaned_data.get('date')
        amount_min = form.cleaned_data.get('amount_min')
        amount_max = form.cleaned_data.get('amount_max')
        contra_type = form.cleaned_data.get('contra_type')
        performed_by = form.cleaned_data.get('performed_by')

        if contra_no:
            contras = contras.filter(contra_no=contra_no)
        if date:
            contras = contras.filter(date=date)
        if amount_min:
            contras = contras.filter(amount__gte=amount_min)
        if amount_max:
            contras = contras.filter(amount__lte=amount_max)
        if contra_type:
            contras = contras.filter(contra_type=contra_type)
        if performed_by:
            contras = contras.filter(performed_by__username=performed_by)

    context = {
        'form': form,
        'contras': contras
    }

    return render(request, 'list_contra.html', context)





import pandas as pd
from django.http import HttpResponse
from .models import Contra

def download_contra_excel(request):
    # Fetch contra transactions
    contras = Contra.objects.all()
    
    # Create a DataFrame
    df = pd.DataFrame(list(contras.values('contra_no', 'date', 'amount', 'contra_type', 'performed_by')))
    
    # Create an HttpResponse object with Excel content type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=contra_transactions.xlsx'
    
    # Write DataFrame to the response
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Contras')
    
    return response



def download_contra_pdf(request):
    # Fetch contra transactions from the database
    contras = Contra.objects.all()  # Add any necessary filtering here

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=contra_transactions.pdf'

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=10, leftMargin=10, topMargin=20, bottomMargin=20)

    elements = []
    data = [
        ['Contra No', 'Date', 'Amount', 'Type', 'Performed By']
    ]

    # Set up a style for paragraphs
    styles = getSampleStyleSheet() # type: ignore
    styleN = styles['BodyText']
    styleN.wordWrap = 'CJK'  # Enable wrapping

    for contra in contras:
        row = [
            Paragraph(str(contra.contra_no) if contra.contra_no else '', styleN),
            Paragraph(contra.date.strftime('%Y-%m-%d') if contra.date else '', styleN),
            Paragraph(str(contra.amount) if contra.amount else '', styleN),
            Paragraph(contra.get_contra_type_display() if contra.contra_type else '', styleN),
            Paragraph(str(contra.performed_by) if contra.performed_by else '', styleN),
        ]
        data.append(row)

    table = Table(data, colWidths=[1.0 * inch] * len(data[0]))
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table.setStyle(style)
    elements.append(table)
    doc.build(elements)

    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response

###############################################################################################################################


from django.shortcuts import render, redirect
from django.contrib import messages
import pandas as pd
from .forms import UploadExcelForm
from .models import Receipt
import pandas as pd

def upload_receipt_from_excel(request):
    if request.method == 'POST':
        form = UploadExcelForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            df = pd.read_excel(excel_file)

            success_count = 0
            fail_count = 0
            failed_rows = []
            failed_reasons = []
            total_rows = len(df)

            for index, row in df.iterrows():
                try:
                    # Parse Receipt Date, support multiple formats
                    receipt_date_str = row.get('Receipt Date')
                    try:
                        receipt_date = pd.to_datetime(receipt_date_str, format='%d %b %Y', errors='raise')
                    except (ValueError, TypeError):
                        receipt_date = pd.to_datetime(receipt_date_str, format='%d/%m/%Y', errors='coerce')

                    mode_of_payment = row.get('Mode of Payment')
                    type_of_receipt = row.get('Type of Receipt')
                    amount = row.get('Amount')

                    # Track missing fields
                    missing_fields = []   
                    if pd.isnull(type_of_receipt):
                        missing_fields.append('Type of Receipt')
                    if pd.isnull(amount):
                        missing_fields.append('Amount')
                    if pd.isnull(receipt_date):
                        missing_fields.append('Receipt Date')

                    # Manual Book No and Manual Receipt No are required for all transactions
                    if pd.isnull(row.get('Manual Book No')):
                        missing_fields.append('Manual Book No')
                    if pd.isnull(row.get('Manual Receipt No')):
                        missing_fields.append('Manual Receipt No')

                    # If mode of payment is Cheque, check for Cheque Number
                    if mode_of_payment == 'Cheque' and pd.isnull(row.get('Cheque Number')):
                        missing_fields.append('Cheque Number')
                    
                    # Get Manual Receipt No; if missing, raise an error
                    manual_receipt_no = row.get('Manual Receipt No')

                    # If any required field is missing, raise an error
                    if missing_fields:
                        raise ValueError(f"Missing required fields: {', '.join(missing_fields)}")

                    # Optional fields handling
                    transaction_id = row.get('Transaction ID') if not pd.isnull(row.get('Transaction ID')) else None
                    cheque_number = row.get('Cheque Number') if not pd.isnull(row.get('Cheque Number')) else None

                    # Create and save the Receipt object
                    receipt = Receipt(
                        manual_book_no=row.get('Manual Book No'),
                        manual_receipt_no=manual_receipt_no,
                        name=row.get('Name'),
                        phone=row.get('Phone'),
                        address=row.get('Address', ''),
                        type_of_receipt=type_of_receipt,
                        mode_of_payment=mode_of_payment,
                        transaction_id=transaction_id,
                        cheque_number=cheque_number,
                        amount=amount,
                        receipt_date=receipt_date
                    )
                    receipt.save()
                    success_count += 1

                except Exception as e:
                    # Collect details for failed rows
                    fail_count += 1
                    failed_rows.append(index + 1)
                    failed_reasons.append(f"Row {index + 1}: {str(e)}")
                
                # Update progress bar for client-side updates (if AJAX or session tracking is used)
                progress = (index + 1) / total_rows * 100
                request.session['upload_progress'] = progress

            # Prepare upload report to display after processing
            report = f"Upload Summary:\n\n"
            report += f"Total Rows Processed: {total_rows}\n"
            report += f"Successful Uploads: {success_count}\n"
            report += f"Failed Uploads: {fail_count}\n\n"
            report += "Failure Details:\n"
            for reason in failed_reasons:
                report += reason + '\n'

            # Store report in session for later display
            request.session['upload_report'] = report

            # Display success and error messages
            messages.success(request, f'Successfully uploaded {success_count} receipts.')
            if fail_count > 0:
                messages.error(request, f'Failed to upload {fail_count} receipts. Check the report for details.')

            # Redirect to the success page to show the report and progress
            return redirect('upload_receipt_success')

    else:
        form = UploadExcelForm()

    return render(request, 'upload_receipt_from_excel.html', {'form': form})



def upload_progress(request):
    """Returns the upload progress."""
    progress = request.session.get('upload_progress', 0)
    return render(request, 'upload_progress.html', {'progress': progress})


def upload_receipt_success(request):
    """Displays the summary report of the upload."""
    report = request.session.get('upload_report', '')
    return render(request, 'upload_receipt_success.html', {'report': report})

###############################################################################################################################




import pandas as pd
from django.contrib.auth.models import User
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Voucher, HeadOfAccount
from .forms import ExcelUploadForm



def upload_vouchers_from_excel(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            df = pd.read_excel(excel_file)

            success_count = 0
            fail_count = 0
            failed_rows = []
            failed_reasons = []
            total_rows = len(df)
            already_exists = 0

            # Get the current user to set as `created_by`
            current_user = request.user

            for index, row in df.iterrows():
                try:
                    # Extract and clean data
                    # Handle multiple date formats
                    try:
                        # First try with 'dd/mm/yyyy' format
                        voucher_date = pd.to_datetime(row.get('Date'), format='%d/%m/%Y', errors='raise')
                    except Exception:
                        try:
                            # Then try with 'dd-mm-yyyy' format
                            voucher_date = pd.to_datetime(row.get('Date'), format='%d-%m-%Y', errors='raise')
                        except Exception:
                            try:
                                # Then try with 'dd Mon yyyy' format (e.g., '21 Mar 2024')
                                voucher_date = pd.to_datetime(row.get('Date'), format='%d %b %Y', errors='raise')
                            except Exception as date_error:
                                raise ValueError(f"Invalid date format in row {index + 1}: {date_error}")

                    mode_of_payment = row.get('Mode Of Payment')
                    transaction_id = row.get('Transaction ID') if not pd.isnull(row.get('Transaction ID')) else None
                    amount = row.get('Amount')
                    
                    # Initialize an empty list to collect missing fields
                    missing_fields = []

                    # Check for missing required fields
                    if pd.isnull(row.get('Voucher No')):
                        missing_fields.append('Voucher No')
                    if pd.isnull(row.get('Paid To')):
                        missing_fields.append('Paid To')
                    if pd.isnull(row.get('On Account Of')):
                        missing_fields.append('On Account Of')
                    if pd.isnull(row.get('Head Of Account')):
                        missing_fields.append('Head Of Account')
                    if pd.isnull(mode_of_payment):
                        missing_fields.append('Mode Of Payment')
                    if pd.isnull(amount):
                        missing_fields.append('Amount')
                    if pd.isnull(voucher_date):
                        missing_fields.append('Date')

                    # Check for 'Approved By' only if status is 'approved'
                    approved_by = row.get('Approved By')
                    status = row.get('Status').strip().lower()  # Normalize status field
                    if status == 'approved':
                        if pd.isnull(approved_by):
                            missing_fields.append('Approved By')
                        else:
                            try:
                                approved_by_user = User.objects.get(username=approved_by.strip())
                            except User.DoesNotExist:
                                raise ValueError(f"User '{approved_by}' does not exist")
                    else:
                        approved_by_user = None

                    if mode_of_payment == 'bank_transfer' or mode_of_payment == 'cheque':
                        if pd.isnull(transaction_id):
                            missing_fields.append('Transaction ID')

                    if missing_fields:
                        raise ValueError(f"Missing required fields: {', '.join(missing_fields)}")

                    # Fetch or create HeadOfAccount instance
                    head_of_account_name = row.get('Head Of Account')
                    head_of_account, created = HeadOfAccount.objects.get_or_create(name=head_of_account_name)

                    # Check for existing vouchers with the same number
                    if Voucher.objects.filter(voucher_no=row.get('Voucher No')).exists():
                        already_exists += 1
                        continue  # Skip to the next row

                    # Create Voucher instance
                    voucher = Voucher(
                        voucher_no=row.get('Voucher No'),
                        paid_to=row.get('Paid To'),
                        on_account_of=row.get('On Account Of'),
                        head_of_account=head_of_account,
                        mode_of_payment=mode_of_payment,
                        transaction_id=transaction_id,
                        amount=amount,
                        voucher_date=voucher_date,
                        received_by=row.get('Received By'),
                        approved_by=approved_by_user,
                        created_by=current_user  # Set created_by to the current user
                    )
                    voucher.save()
                    success_count += 1
                except Exception as e:
                    fail_count += 1
                    failed_rows.append(index)
                    failed_reasons.append(f"Row {index + 1}: {str(e)}")

                # Update progress bar (assuming an AJAX call to update progress on the client side)
                progress = (index + 1) / total_rows * 100
                request.session['upload_progress'] = progress

            # Prepare upload report
            report = f"Upload Summary:\n\n"
            report += f"Total Rows Processed: {total_rows}\n"
            report += f"Successful Uploads: {success_count}\n"
            report += f"Failed Uploads: {fail_count}\n"
            report += f"Already Exists: {already_exists}\n\n"
            report += "Failure Details:\n"
            for reason in failed_reasons:
                report += reason + '\n'

            # Store report in session
            request.session['upload_report'] = report

            # Display success and error messages
            messages.success(request, f'Successfully uploaded {success_count} vouchers.')
            if fail_count > 0:
                messages.error(request, f'Failed to upload {fail_count} vouchers.')

            return redirect('upload_vouchers_success')
    else:
        form = ExcelUploadForm()
    return render(request, 'upload_vouchers_from_excel.html', {'form': form})



import pandas as pd
from django.contrib.auth.models import User
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Voucher, HeadOfAccount
from .forms import ExcelUploadForm

def upload_vouchers_from_excel(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            df = pd.read_excel(excel_file)

            success_count = 0
            fail_count = 0
            failed_rows = []
            failed_reasons = []
            total_rows = len(df)
            already_exists = 0

            # Get the current user to set as `created_by`
            current_user = request.user

            for index, row in df.iterrows():
                try:
                    # Extract and clean data
                    voucher_date = pd.to_datetime(row.get('Date'), errors='coerce')
                    mode_of_payment = row.get('Mode Of Payment')
                    transaction_id = row.get('Transaction ID') if not pd.isnull(row.get('Transaction ID')) else None
                    amount = row.get('Amount')
                    
                    # Initialize an empty list to collect missing fields
                    missing_fields = []

                    # Check for missing required fields
                    if pd.isnull(row.get('Voucher No')):
                        missing_fields.append('Voucher No')
                    if pd.isnull(row.get('Paid To')):
                        missing_fields.append('Paid To')
                    if pd.isnull(row.get('On Account Of')):
                        missing_fields.append('On Account Of')
                    if pd.isnull(row.get('Head Of Account')):
                        missing_fields.append('Head Of Account')
                    if pd.isnull(mode_of_payment):
                        missing_fields.append('Mode Of Payment')
                    if pd.isnull(amount):
                        missing_fields.append('Amount')
                    if pd.isnull(voucher_date):
                        missing_fields.append('Date')

                    # Check for 'Approved By' only if status is 'approved'
                    approved_by = row.get('Approved By')
                    status = row.get('Status').strip().lower()  # Normalize status field

                    if status == 'approved':
                        if pd.isnull(approved_by):
                            missing_fields.append('Approved By')
                        else:
                            try:
                                approved_by_user = User.objects.get(username=approved_by.strip())
                            except User.DoesNotExist:
                                raise ValueError(f"User '{approved_by}' does not exist")
                    else:
                        approved_by_user = None

                    if mode_of_payment == 'bank_transfer' or mode_of_payment == 'cheque':
                        if pd.isnull(transaction_id):
                            missing_fields.append('Transaction ID')

                    if missing_fields:
                        raise ValueError(f"Missing required fields: {', '.join(missing_fields)}")

                    # Fetch or create HeadOfAccount instance
                    head_of_account_name = row.get('Head Of Account')
                    head_of_account, created = HeadOfAccount.objects.get_or_create(name=head_of_account_name)

                    # Check for existing vouchers with the same number
                    if Voucher.objects.filter(voucher_no=row.get('Voucher No')).exists():
                        already_exists += 1
                        continue  # Skip to the next row

                    # Create Voucher instance
                    voucher = Voucher(
                        voucher_no=row.get('Voucher No'),
                        paid_to=row.get('Paid To'),
                        on_account_of=row.get('On Account Of'),
                        head_of_account=head_of_account,
                        mode_of_payment=mode_of_payment,
                        transaction_id=transaction_id,
                        amount=amount,
                        voucher_date=voucher_date,
                        received_by=row.get('Received By'),
                        approved_by=approved_by_user,
                        created_by=current_user  # Set created_by to the current user
                    )
                    # Set the status explicitly
                    voucher.status = status  # Update status field
                    voucher.save()
                    success_count += 1
                except Exception as e:
                    fail_count += 1
                    failed_rows.append(index)
                    failed_reasons.append(f"Row {index + 1}: {str(e)}")

                # Update progress bar (assuming an AJAX call to update progress on the client side)
                progress = (index + 1) / total_rows * 100
                request.session['upload_progress'] = progress

            # Prepare upload report
            report = f"Upload Summary:\n\n"
            report += f"Total Rows Processed: {total_rows}\n"
            report += f"Successful Uploads: {success_count}\n"
            report += f"Failed Uploads: {fail_count}\n"
            report += f"Already Exists: {already_exists}\n\n"
            report += "Failure Details:\n"
            for reason in failed_reasons:
                report += reason + '\n'

            # Store report in session
            request.session['upload_report'] = report

            # Display success and error messages
            messages.success(request, f'Successfully uploaded {success_count} vouchers.')
            if fail_count > 0:
                messages.error(request, f'Failed to upload {fail_count} vouchers.')

            return redirect('upload_vouchers_success')
    else:
        form = ExcelUploadForm()
    return render(request, 'upload_vouchers_from_excel.html', {'form': form})

def upload_vouchers_success(request):
    return render(request, 'upload_vouchers_success.html')

##################################################

from django.shortcuts import render
from django.db.models import Sum, Count
from .models import Receipt
from django.db.models import Count, Sum
from django.shortcuts import render
from .models import Receipt  # Assuming Receipt model contains necessary fields
from django.db.models import Count, Sum
from django.shortcuts import render
from .models import Receipt  # Assuming Receipt model contains necessary fields

def book_summary(request):
    # Aggregate data by manual_book_no (book number) and mode_of_payment
    book_summary_data = Receipt.objects.values(
        'manual_book_no',  # Grouping by book number
        'mode_of_payment'  # Grouping by mode of payment
    ).annotate(
        receipt_count=Count('id'),  # Counting the number of receipts for each group
        total_amount=Sum('amount')  # Summing the total amount for each group
    )
    
    # Initialize an empty dictionary to store the data for each book number
    summary = {}

    # Loop over the aggregated data to organize it by book number
    for entry in book_summary_data:
        book_number = entry['manual_book_no']
        
        # Initialize book entry if not already present
        if book_number not in summary:
            summary[book_number] = {
                'book_number': book_number,
                'receipt_count': 0,
                'cash': 0,
                'upi': 0,
                'cheque': 0,
                'bank_transfer': 0,
                'total_amount': 0
            }
        
        # Add the aggregated data to the respective fields
        summary[book_number]['receipt_count'] += entry['receipt_count']
        summary[book_number]['total_amount'] += entry['total_amount']
        
        # Assign amounts based on mode of payment
        if entry['mode_of_payment'] == 'Cash':
            summary[book_number]['cash'] += entry['total_amount']
        elif entry['mode_of_payment'] == 'UPI':
            summary[book_number]['upi'] += entry['total_amount']
        elif entry['mode_of_payment'] == 'Cheque':
            summary[book_number]['cheque'] += entry['total_amount']
        elif entry['mode_of_payment'] == 'Bank Transfer':
            summary[book_number]['bank_transfer'] += entry['total_amount']
    
    # Convert the summary dictionary to a list for passing to the template
    summary_data = list(summary.values())

    # Return the data to the template
    return render(request, 'book_summary.html', {
        'book_summary': summary_data
    })


#################################################################################################################


